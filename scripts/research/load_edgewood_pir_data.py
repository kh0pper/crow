#!/usr/bin/env python3
"""
Load Edgewood ISD ORR #129 response data into the TEA research database.

Parses campus-level per-pupil expenditures (19 xlsx files), Title I and
Compensatory Education allocations (1 xlsx, 2 sheets), and a Facility
Condition Assessment with 441 deferred maintenance line items (1 xlsx).

Tables:
  - research_edgewood_campus_financials (new)
  - research_fca_cost_items (new)

Usage:
    python scripts/load_edgewood_pir_data.py --all
    python scripts/load_edgewood_pir_data.py --financials
    python scripts/load_edgewood_pir_data.py --fca
    python scripts/load_edgewood_pir_data.py --verify
"""

import argparse
import glob
import os
import re
import shutil
import sqlite3
import sys
from datetime import datetime

import openpyxl

DB_PATH = os.path.join(
    os.path.dirname(__file__), "..", "texas-gov-data-mcp", "data", "tea_data.db"
)
SOURCES_DIR = os.path.join(
    os.path.dirname(__file__), "..", "insd-5941", "sources", "edgewood-isd"
)
DISTRICT_TEA_ID = "015905"
SOURCE_LABEL = "ORR #129 - Edgewood ISD"


def backup_db():
    """Create a timestamped backup of the database."""
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = DB_PATH + f".bak-{ts}"
    shutil.copy2(DB_PATH, backup)
    print(f"  Backup: {backup}")


def create_tables(conn):
    """Create research tables if they don't exist. Idempotent."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS research_edgewood_campus_financials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_tea_id TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            school_year TEXT NOT NULL,
            fund_category TEXT NOT NULL,
            enrollment INTEGER,
            budget REAL,
            expenditure REAL,
            per_pupil REAL,
            notes TEXT,
            UNIQUE(source, campus_tea_id, school_year, fund_category)
        );

        CREATE TABLE IF NOT EXISTS research_fca_cost_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            district_tea_id TEXT NOT NULL,
            campus_or_facility TEXT,
            category TEXT,
            description TEXT,
            priority INTEGER,
            original_cost REAL,
            revised_cost REAL,
            cost_notes TEXT,
            row_seq INTEGER NOT NULL,
            UNIQUE(source, district_tea_id, row_seq)
        );
        CREATE INDEX IF NOT EXISTS ix_rfci_district
            ON research_fca_cost_items(district_tea_id);
    """)
    print("  Tables verified/created.")


# ── Cost parsing ────────────────────────────────────────────────────────────


def parse_cost(value):
    """Parse a cost cell value into (cost_float, cost_notes_str).

    Returns:
        (float, None) for clean numeric values
        (float, str) for values like "50000 each" where a number can be extracted
        (None, str) for unit rates like "250/sf" or references like "see above"
        (None, None) for None/empty
    """
    if value is None:
        return None, None

    # Numeric values pass through directly
    if isinstance(value, (int, float)):
        return float(value), None

    s = str(value).strip()
    if not s:
        return None, None

    # Try clean numeric first (handles "50,000" strings)
    cleaned = s.replace(",", "").replace("$", "").strip()
    try:
        return float(cleaned), None
    except ValueError:
        pass

    # "50000 each" or "50,000 each" — extract the number, note "unit cost"
    m = re.match(r"^[\$]?\s*([\d,]+)\s+each\b", s, re.IGNORECASE)
    if m:
        num = float(m.group(1).replace(",", ""))
        return num, "unit cost"

    # Unit rates like "500/light", "1200/door", "250/sf" — store as note only
    if "/" in s and re.search(r"\d", s):
        return None, s

    # Reference values: "with roof", "see above", "need info", "?", "ranges", "above"
    return None, s


# ── PPE (Per-Pupil Expenditures) ────────────────────────────────────────────


def parse_ppe_files():
    """Parse all PPE xlsx files.

    Each file has campus identity in row 1 header:
        'Per Pupil Expenditures All Funds for 2024 - 2025 for (015905002) - John F Kennedy H S'
    Row 8 col B = enrollment, row 9 col B = per_pupil, row 10 col B = total cost.
    """
    pattern = os.path.join(SOURCES_DIR, "Per Pupil*.xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        print("  WARNING: No PPE xlsx files found")
        return []

    header_re = re.compile(r"\((\d+)\)\s*-\s*(.*)")
    rows = []
    seen_ids = set()

    for fpath in files:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        header = ws.cell(row=1, column=1).value
        if not header:
            print(f"  WARNING: Empty header in {os.path.basename(fpath)}")
            continue

        m = header_re.search(header)
        if not m:
            print(f"  WARNING: Cannot parse header: {header}")
            continue

        campus_tea_id = m.group(1)
        campus_name = m.group(2).strip()

        if campus_tea_id in seen_ids:
            print(f"  WARNING: Duplicate campus {campus_tea_id} in {os.path.basename(fpath)}")
            continue
        seen_ids.add(campus_tea_id)

        enrollment = ws.cell(row=8, column=2).value
        per_pupil = ws.cell(row=9, column=2).value
        total_cost = ws.cell(row=10, column=2).value

        rows.append({
            "source": SOURCE_LABEL,
            "campus_tea_id": campus_tea_id,
            "campus_name": campus_name,
            "school_year": "2024-2025",
            "fund_category": "ppe_all_funds",
            "enrollment": int(enrollment) if enrollment else None,
            "budget": None,
            "expenditure": float(total_cost) if total_cost else None,
            "per_pupil": float(per_pupil) if per_pupil else None,
            "notes": None,
        })

    print(f"  PPE: parsed {len(rows)} campuses from {len(files)} files")
    return rows


# ── Title I & Comp Ed ───────────────────────────────────────────────────────


def parse_title_i_comp_ed():
    """Parse Title I and Compensatory Education from the combined xlsx.

    Title I sheet: Fund(A), Org(B), Campus(C), Year(D), Budget(E), Transactions(F)
    Comp Ed sheet: Fund(A), Org(B), PIC(C), Campus(D), Year(E), Budget(F), Transactions(G)

    Org code maps to campus TEA ID: 015905 + zero-padded 3-digit org.
    """
    fpath = os.path.join(SOURCES_DIR, "Title I_State Comp_Allocations and Expenditures.xlsx")
    if not os.path.exists(fpath):
        print(f"  WARNING: Title I/Comp Ed file not found")
        return []

    wb = openpyxl.load_workbook(fpath, data_only=True)
    rows = []

    # Title I sheet
    if "Title I" in wb.sheetnames:
        ws = wb["Title I"]
        for row_idx in range(2, ws.max_row + 1):
            org = ws.cell(row=row_idx, column=2).value  # B = Org
            campus_name = ws.cell(row=row_idx, column=3).value  # C = Campus
            budget = ws.cell(row=row_idx, column=5).value  # E = Budget
            txn = ws.cell(row=row_idx, column=6).value  # F = Transactions

            if org is None or campus_name is None:
                continue

            campus_tea_id = f"{DISTRICT_TEA_ID}{int(org):03d}"

            rows.append({
                "source": SOURCE_LABEL,
                "campus_tea_id": campus_tea_id,
                "campus_name": str(campus_name).strip(),
                "school_year": "2025-2026",
                "fund_category": "title_i",
                "enrollment": None,
                "budget": float(budget) if budget else None,
                "expenditure": float(txn) if txn else None,
                "per_pupil": None,
                "notes": "Fund 211",
            })
        print(f"  Title I: parsed {sum(1 for r in rows if r['fund_category'] == 'title_i')} rows")

    # Compensatory Education sheet
    if "Compensatory Education" in wb.sheetnames:
        ws = wb["Compensatory Education"]
        comp_count = 0
        for row_idx in range(2, ws.max_row + 1):
            org = ws.cell(row=row_idx, column=2).value  # B = Org
            campus_name = ws.cell(row=row_idx, column=4).value  # D = Campus
            budget = ws.cell(row=row_idx, column=6).value  # F = Budget
            txn = ws.cell(row=row_idx, column=7).value  # G = Transactions

            if org is None or campus_name is None:
                continue

            campus_tea_id = f"{DISTRICT_TEA_ID}{int(org):03d}"

            rows.append({
                "source": SOURCE_LABEL,
                "campus_tea_id": campus_tea_id,
                "campus_name": str(campus_name).strip(),
                "school_year": "2025-2026",
                "fund_category": "comp_ed",
                "enrollment": None,
                "budget": float(budget) if budget else None,
                "expenditure": float(txn) if txn else None,
                "per_pupil": None,
                "notes": "Fund 199 PIC 30",
            })
            comp_count += 1
        print(f"  Comp Ed: parsed {comp_count} rows")

    return rows


# ── FCA (Facility Condition Assessment) ─────────────────────────────────────


def parse_fca():
    """Parse FCA deferred maintenance line items.

    Columns: Category(A), Campus(B), Description(C), Priority(D),
             OriginalCost(E), RevisedCost(F), CostNotes(G)

    Skips fully empty rows. Loads subtotal/continuation rows (NULL campus).
    """
    fpath = os.path.join(SOURCES_DIR, "FCA.xlsx")
    if not os.path.exists(fpath):
        print(f"  WARNING: FCA.xlsx not found")
        return []

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = []
    row_seq = 0

    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 8)]

        # Skip fully empty rows
        if all(v is None for v in vals):
            continue

        row_seq += 1

        category = vals[0]
        campus = str(vals[1]).strip() if vals[1] else None
        description = str(vals[2]).strip() if vals[2] else None
        raw_priority = vals[3]
        raw_original = vals[4]
        raw_revised = vals[5]
        cost_notes_cell = str(vals[6]).strip() if vals[6] else None

        # Priority: store as integer if whole number, else None (averages in subtotal rows)
        priority = None
        if raw_priority is not None:
            try:
                p = float(raw_priority)
                if p == int(p):
                    priority = int(p)
            except (ValueError, TypeError):
                pass

        # Parse costs
        orig_cost, orig_note = parse_cost(raw_original)
        rev_cost, rev_note = parse_cost(raw_revised)

        # Merge cost notes: cell note + any parsing notes
        notes_parts = []
        if cost_notes_cell:
            notes_parts.append(cost_notes_cell)
        if orig_note:
            notes_parts.append(f"original: {orig_note}")
        if rev_note:
            notes_parts.append(f"revised: {rev_note}")
        combined_notes = "; ".join(notes_parts) if notes_parts else None

        rows.append({
            "source": SOURCE_LABEL,
            "district_tea_id": DISTRICT_TEA_ID,
            "campus_or_facility": campus,
            "category": str(category).strip() if category else None,
            "description": description,
            "priority": priority,
            "original_cost": orig_cost,
            "revised_cost": rev_cost,
            "cost_notes": combined_notes,
            "row_seq": row_seq,
        })

    print(f"  FCA: parsed {len(rows)} rows (skipped {ws.max_row - 1 - len(rows)} empty)")
    return rows


# ── Load functions ──────────────────────────────────────────────────────────


def load_financials(conn):
    """Load PPE + Title I + Comp Ed into research_edgewood_campus_financials."""
    ppe_rows = parse_ppe_files()
    ti_ce_rows = parse_title_i_comp_ed()
    all_rows = ppe_rows + ti_ce_rows

    total = 0
    for row in all_rows:
        conn.execute("""
            INSERT OR REPLACE INTO research_edgewood_campus_financials
            (source, campus_tea_id, campus_name, school_year, fund_category,
             enrollment, budget, expenditure, per_pupil, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row["source"], row["campus_tea_id"], row["campus_name"],
            row["school_year"], row["fund_category"],
            row["enrollment"], row["budget"], row["expenditure"],
            row["per_pupil"], row["notes"],
        ))
        total += 1

    conn.commit()
    print(f"  Financials total: inserted/updated {total} rows")
    return total


def load_fca(conn):
    """Load FCA into research_fca_cost_items."""
    rows = parse_fca()

    total = 0
    for row in rows:
        conn.execute("""
            INSERT OR REPLACE INTO research_fca_cost_items
            (source, district_tea_id, campus_or_facility, category, description,
             priority, original_cost, revised_cost, cost_notes, row_seq)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row["source"], row["district_tea_id"], row["campus_or_facility"],
            row["category"], row["description"],
            row["priority"], row["original_cost"], row["revised_cost"],
            row["cost_notes"], row["row_seq"],
        ))
        total += 1

    conn.commit()
    print(f"  FCA total: inserted/updated {total} rows")
    return total


# ── Verification ────────────────────────────────────────────────────────────


def verify(conn):
    """Verify loaded data against expected values."""
    print("\n=== Verification ===\n")
    errors = []

    # 1. Financials by category
    print("Financials by fund_category:")
    cursor = conn.execute("""
        SELECT fund_category, COUNT(*)
        FROM research_edgewood_campus_financials
        WHERE source = ?
        GROUP BY fund_category
        ORDER BY fund_category
    """, (SOURCE_LABEL,))
    counts = {}
    for row in cursor:
        counts[row[0]] = row[1]
        print(f"  {row[0]}: {row[1]} rows")

    expected = {"ppe_all_funds": 19, "title_i": 17, "comp_ed": 17}
    for cat, exp in expected.items():
        actual = counts.get(cat, 0)
        if actual != exp:
            errors.append(f"Financials {cat}: expected {exp}, got {actual}")

    # 2. FCA row count
    cursor = conn.execute("""
        SELECT COUNT(*) FROM research_fca_cost_items
        WHERE district_tea_id = ?
    """, (DISTRICT_TEA_ID,))
    fca_count = cursor.fetchone()[0]
    print(f"\nFCA rows: {fca_count}")
    if fca_count != 441:
        errors.append(f"FCA: expected 441, got {fca_count}")

    # 3. Spot check: Kennedy HS PPE
    print("\nSpot checks:")
    cursor = conn.execute("""
        SELECT campus_name, enrollment, per_pupil, expenditure
        FROM research_edgewood_campus_financials
        WHERE campus_tea_id = '015905002' AND fund_category = 'ppe_all_funds'
    """)
    row = cursor.fetchone()
    if row:
        print(f"  Kennedy HS PPE: enroll={row[1]}, ppe=${row[2]:,.2f}, total=${row[3]:,.0f}")
        if row[1] != 998:
            errors.append(f"Kennedy enrollment: expected 998, got {row[1]}")
        if abs(row[2] - 8032.64) > 0.01:
            errors.append(f"Kennedy PPE: expected 8032.64, got {row[2]:.2f}")
    else:
        errors.append("Kennedy HS PPE row not found")

    # 4. Spot check: Loma Park Comp Ed
    cursor = conn.execute("""
        SELECT campus_name, budget, expenditure
        FROM research_edgewood_campus_financials
        WHERE campus_tea_id = '015905112' AND fund_category = 'comp_ed'
    """)
    row = cursor.fetchone()
    if row:
        print(f"  Loma Park Comp Ed: budget=${row[1]:,.0f}, expenditure=${row[2]:,.2f}")
        if abs(row[1] - 210220) > 1:
            errors.append(f"Loma Park budget: expected 210220, got {row[1]}")
        if abs(row[2] - 120377.45) > 0.01:
            errors.append(f"Loma Park expenditure: expected 120377.45, got {row[2]:.2f}")
    else:
        errors.append("Loma Park Comp Ed row not found")

    # 5. Spot check: FCA Memorial HS revised costs
    cursor = conn.execute("""
        SELECT SUM(revised_cost)
        FROM research_fca_cost_items
        WHERE campus_or_facility = 'Memorial High School'
    """)
    row = cursor.fetchone()
    if row and row[0]:
        print(f"  FCA Memorial HS revised total: ${row[0]:,.0f}")
        if abs(row[0] - 7124200) > 1:
            errors.append(f"FCA Memorial revised: expected 7124200, got {row[0]}")
    else:
        errors.append("FCA Memorial HS rows not found or all NULL costs")

    # 6. Summary stats
    print("\nSummary:")
    cursor = conn.execute("""
        SELECT SUM(budget), SUM(expenditure)
        FROM research_edgewood_campus_financials
        WHERE fund_category = 'comp_ed'
    """)
    row = cursor.fetchone()
    if row:
        print(f"  Comp Ed total: budget=${row[0]:,.0f}, spent=${row[1]:,.2f} "
              f"({row[1]/row[0]*100:.1f}% YTD)")

    cursor = conn.execute("""
        SELECT SUM(budget), SUM(expenditure)
        FROM research_edgewood_campus_financials
        WHERE fund_category = 'title_i'
    """)
    row = cursor.fetchone()
    if row:
        print(f"  Title I total: budget=${row[0]:,.0f}, spent=${row[1]:,.2f} "
              f"({row[1]/row[0]*100:.1f}% YTD)")

    cursor = conn.execute("""
        SELECT SUM(original_cost), SUM(revised_cost)
        FROM research_fca_cost_items
        WHERE district_tea_id = ?
    """, (DISTRICT_TEA_ID,))
    row = cursor.fetchone()
    if row:
        orig = row[0] or 0
        rev = row[1] or 0
        print(f"  FCA totals: original=${orig:,.0f}, revised=${rev:,.0f}")

    # Report
    if errors:
        print(f"\n*** {len(errors)} VERIFICATION ERROR(S) ***")
        for e in errors:
            print(f"  - {e}")
        return False
    else:
        print("\n  All checks passed.")
        return True


# ── Main ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Load Edgewood ISD ORR #129 data")
    parser.add_argument("--all", action="store_true", help="Load all data types")
    parser.add_argument("--financials", action="store_true", help="Load PPE + Title I + Comp Ed")
    parser.add_argument("--fca", action="store_true", help="Load FCA deferred maintenance")
    parser.add_argument("--verify", action="store_true", help="Verify loaded data")

    args = parser.parse_args()

    if not any([args.all, args.financials, args.fca, args.verify]):
        parser.print_help()
        sys.exit(1)

    print(f"Database: {os.path.abspath(DB_PATH)}")
    conn = sqlite3.connect(DB_PATH)

    try:
        if args.all or args.financials or args.fca:
            backup_db()
            create_tables(conn)

        if args.all or args.financials:
            print("\n--- Campus Financials ---")
            load_financials(conn)

        if args.all or args.fca:
            print("\n--- Facility Condition Assessment ---")
            load_fca(conn)

        if args.verify or args.all:
            ok = verify(conn)
            if not ok:
                sys.exit(1)

    finally:
        conn.close()


if __name__ == "__main__":
    main()
