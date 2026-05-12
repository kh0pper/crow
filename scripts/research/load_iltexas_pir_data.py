#!/usr/bin/env python3
"""
Load ILTexas Supplemental PIR Response Data (Mar 2026)

Sources:
- PIR - total teachers by org.xlsx (teacher headcounts by campus, 4 years)
- PIR 2.11.2026.xlsx (salary schedules, staffing details, Liberty County data)
- PEIMS campus financial PDFs (per-campus expenditure reports, 2021-2025)
- ILTexas Liberty County Construction Projects Summary 2.25.csv

Target: texas-gov-data-mcp/data/tea_data.db
Tables: research_charter_campus_summary, research_charter_financial
"""

import glob
import os
import re
import sqlite3
import subprocess
import sys

import openpyxl

CAPSTONE_ROOT = os.environ.get(
    "CAPSTONE_ROOT", os.path.expanduser("~/spring-2026")
)
BASE_DIR = os.environ.get(
    "ILTEXAS_PIR_DIR",
    os.path.join(CAPSTONE_ROOT, "insd-5941", "sources", "iltexas-pir-supplemental"),
)
DB_PATH = os.environ.get(
    "TEA_DB_PATH",
    os.path.join(CAPSTONE_ROOT, "texas-gov-data-mcp", "data", "tea_data.db"),
)
SOURCE = "PIR #11 - ILTexas Supplemental (Mar 2026)"
DISTRICT_TEA_ID = "057848"


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def create_tables(conn):
    """Create tables for ILTexas PIR data."""

    # PEIMS campus financial data (per-campus expenditure by function)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_peims_campus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_tea_id TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            school_year TEXT NOT NULL,
            data_type TEXT NOT NULL,  -- 'actual' or 'budgeted'
            enrollment INTEGER,
            category TEXT NOT NULL,   -- 'object', 'function', or 'program'
            line_item TEXT NOT NULL,
            general_fund_amount REAL,
            general_fund_pct REAL,
            general_fund_per_student REAL,
            all_funds_amount REAL,
            all_funds_pct REAL,
            all_funds_per_student REAL,
            notes TEXT,
            UNIQUE(campus_tea_id, school_year, category, line_item)
        )
    """)

    # Teacher headcounts by campus
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_teacher_counts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            school_year TEXT NOT NULL,
            teacher_count INTEGER NOT NULL,
            notes TEXT,
            UNIQUE(campus_name, school_year)
        )
    """)

    # Salary schedules
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_salary_schedule (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            fiscal_year TEXT NOT NULL,
            posting_region TEXT NOT NULL,
            step INTEGER NOT NULL,
            non_certified_salary REAL NOT NULL,
            certified_salary REAL NOT NULL,
            UNIQUE(fiscal_year, posting_region, step)
        )
    """)

    # Campus average salaries
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_campus_avg_salary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            fiscal_year TEXT NOT NULL,
            avg_teacher_salary REAL,
            UNIQUE(campus_name, fiscal_year)
        )
    """)

    # TA/para staffing by campus
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_staffing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            school_year TEXT NOT NULL,
            total_ta_paras INTEGER,
            ta_to_grade_level INTEGER,
            total_certified_teachers INTEGER,
            total_support_staff_ftes REAL,
            notes TEXT,
            UNIQUE(campus_name, school_year)
        )
    """)

    # Construction costs
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_iltexas_construction (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            campus_name TEXT NOT NULL,
            construction_cost REAL,
            general_contractor TEXT,
            architect TEXT,
            sq_ft INTEGER,
            design_enrollment INTEGER,
            arch_engineering_cost REAL,
            ffe_cost REAL,
            site_work_cost REAL,
            inspectors_cost REAL,
            land_acquisition_cost REAL,
            capital_budget TEXT,
            notes TEXT,
            UNIQUE(campus_name)
        )
    """)

    conn.commit()


def parse_currency(val):
    """Parse currency string like '$2,418,141' or '$-12' to float."""
    if not val or val.strip() == '$0' or val.strip() == '':
        return 0.0
    val = val.strip().replace('$', '').replace(',', '')
    try:
        return float(val)
    except ValueError:
        return None


def parse_pct(val):
    """Parse percentage string like '100.00%' to float."""
    if not val or val.strip() == '':
        return None
    val = val.strip().replace('%', '')
    try:
        return float(val)
    except ValueError:
        return None


def parse_peims_pdf(pdf_path):
    """Parse a PEIMS campus financial PDF and return structured data."""
    result = subprocess.run(
        ['pdftotext', '-layout', pdf_path, '-'],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  ERROR: pdftotext failed for {pdf_path}")
        return None

    text = result.stdout
    lines = text.split('\n')

    # Extract header info
    campus_tea_id = None
    campus_name = None
    school_year = None
    enrollment = None
    data_type = None  # 'actual' or 'budgeted'

    for line in lines[:10]:
        # School year and type
        m = re.search(r'(\d{4}-\d{4}) PEIMS (ACTUAL|BUDGETED) FINANCIAL DATA', line)
        if m:
            school_year = m.group(1)
            data_type = m.group(2).lower()

        # Campus name and TEA ID
        m = re.search(r'TOTALS FOR (.+?)\s*\((\d{9})\)', line)
        if m:
            campus_name = m.group(1).strip()
            campus_tea_id = m.group(2)

        # Enrollment
        m = re.search(r'Total Enrolled Membership:\s*([\d,]+)', line)
        if m:
            enrollment = int(m.group(1).replace(',', ''))

    if not all([campus_tea_id, campus_name, school_year]):
        print(f"  WARN: Could not parse header from {os.path.basename(pdf_path)}")
        return None

    records = []

    # Parse expenditure lines using regex
    # Pattern: label followed by dollar amounts and percentages
    # General Fund amount, %, per student, All Funds amount, %, per student
    expenditure_pattern = re.compile(
        r'^(.+?)\s+'
        r'\$([\d,.-]+)\s+'    # gen fund amount
        r'([\d.]+)%\s+'       # gen fund %
        r'\$([\d,.-]+)\s+'    # gen fund per student
        r'\$([\d,.-]+)\s+'    # all funds amount
        r'([\d.]+)%\s+'       # all funds %
        r'\$([\d,.-]+)'       # all funds per student
    )

    current_category = None

    for line in lines:
        line_stripped = line.strip()

        # Detect category
        if 'Expenditures by Object' in line:
            current_category = 'object'
            continue
        elif 'Expenditures by Function' in line:
            current_category = 'function'
            continue
        elif 'Program expenditures by Program' in line or 'expenditures by Program' in line:
            current_category = 'program'
            continue

        if current_category is None:
            continue

        # Try to match expenditure line
        m = expenditure_pattern.search(line)
        if m:
            label = m.group(1).strip().rstrip('*')
            label = re.sub(r'\s+', ' ', label).strip()

            # Skip section headers that got matched
            if label in ('General', 'All'):
                continue

            records.append({
                'campus_tea_id': campus_tea_id,
                'campus_name': campus_name,
                'school_year': school_year,
                'data_type': data_type,
                'enrollment': enrollment,
                'category': current_category,
                'line_item': label,
                'general_fund_amount': parse_currency('$' + m.group(2)),
                'general_fund_pct': parse_pct(m.group(3) + '%'),
                'general_fund_per_student': parse_currency('$' + m.group(4)),
                'all_funds_amount': parse_currency('$' + m.group(5)),
                'all_funds_pct': parse_pct(m.group(6) + '%'),
                'all_funds_per_student': parse_currency('$' + m.group(7)),
            })

    return records


def load_peims_campus_data(conn):
    """Parse all PEIMS campus financial PDFs and load to DB."""
    print("\n=== Loading PEIMS Campus Financial Data ===")

    # Collect all PEIMS PDF paths
    pdf_dirs = [
        BASE_DIR,                          # 2021-22, 2022-23 loose PDFs
        os.path.join(BASE_DIR, 'peims_zip1'),  # 2022-23, 2023-24, 2024-25
        os.path.join(BASE_DIR, 'peims_zip2'),  # 2023-24
    ]

    pdf_files = []
    for d in pdf_dirs:
        pdf_files.extend(glob.glob(os.path.join(d, '*PEIMS_Campus_Data*.pdf')))

    # Deduplicate by filename (zips may overlap with loose files)
    seen = {}
    for f in pdf_files:
        basename = os.path.basename(f)
        # Prefer zip versions (they have cleaner names)
        if basename not in seen or 'zip' in f:
            seen[basename] = f
    pdf_files = sorted(seen.values())

    print(f"Found {len(pdf_files)} PEIMS PDFs to process")

    total_records = 0
    total_campuses = 0
    errors = 0

    for pdf_path in pdf_files:
        records = parse_peims_pdf(pdf_path)
        if records is None:
            errors += 1
            continue

        total_campuses += 1
        for r in records:
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO research_iltexas_peims_campus
                    (source, campus_tea_id, campus_name, school_year, data_type,
                     enrollment, category, line_item,
                     general_fund_amount, general_fund_pct, general_fund_per_student,
                     all_funds_amount, all_funds_pct, all_funds_per_student)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    SOURCE, r['campus_tea_id'], r['campus_name'], r['school_year'],
                    r['data_type'], r['enrollment'], r['category'], r['line_item'],
                    r['general_fund_amount'], r['general_fund_pct'],
                    r['general_fund_per_student'], r['all_funds_amount'],
                    r['all_funds_pct'], r['all_funds_per_student']
                ))
                total_records += 1
            except sqlite3.IntegrityError as e:
                pass  # Duplicate, skip

    conn.commit()
    print(f"Loaded {total_records} records from {total_campuses} campus reports ({errors} parse errors)")
    return total_records


def load_teacher_counts(conn):
    """Load teacher headcounts from PIR - total teachers by org.xlsx"""
    print("\n=== Loading Teacher Counts ===")

    wb = openpyxl.load_workbook(
        os.path.join(BASE_DIR, 'PIR - total teachers by org .xlsx'),
        data_only=True
    )

    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Extract school year from sheet name (e.g., "2021-2022")
        school_year = sheet_name.strip()

        for row in ws.iter_rows(min_row=2, values_only=False):
            campus_name = str(row[0].value).strip() if row[0].value else None
            count = row[1].value if len(row) > 1 and row[1].value else None

            if campus_name and count and campus_name != 'Position':
                # Clean campus name (remove trailing colon)
                campus_name = campus_name.rstrip(':')
                try:
                    conn.execute("""
                        INSERT OR REPLACE INTO research_iltexas_teacher_counts
                        (source, campus_name, school_year, teacher_count)
                        VALUES (?, ?, ?, ?)
                    """, (SOURCE, campus_name, school_year, int(count)))
                    total += 1
                except (ValueError, sqlite3.IntegrityError):
                    pass

    conn.commit()
    print(f"Loaded {total} teacher count records")
    return total


def load_salary_schedules(conn):
    """Load salary schedules from PIR 2.11.2026.xlsx"""
    print("\n=== Loading Salary Schedules ===")

    wb = openpyxl.load_workbook(
        os.path.join(BASE_DIR, 'PIR 2.11.2026.xlsx'),
        data_only=True
    )

    ws = wb['5. Teacher Salary Schedules and']
    rows = list(ws.iter_rows(values_only=True))

    # FY23 cols 0-1, FY24 cols 3-4, FY25 cols 6-7
    fy_configs = [
        ('FY23', 'HOUSTON', 0, 1),
        ('FY24', 'HOUSTON', 3, 4),
        ('FY25', 'HOUSTON', 6, 7),
    ]

    salary_total = 0
    avg_total = 0

    for fy, region, nc_col, c_col in fy_configs:
        step = 0
        for row in rows[3:]:  # Skip header rows
            nc_val = row[nc_col]
            c_val = row[c_col]

            if nc_val and c_val and isinstance(nc_val, (int, float)) and isinstance(c_val, (int, float)):
                # Check if this is salary data (> 10000) vs avg salary
                if nc_val > 10000 and c_val > 10000:
                    step += 1
                    conn.execute("""
                        INSERT OR REPLACE INTO research_iltexas_salary_schedule
                        (source, fiscal_year, posting_region, step, non_certified_salary, certified_salary)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (SOURCE, fy, region, step, float(nc_val), float(c_val)))
                    salary_total += 1

        # Average salary rows (at bottom)
        for row in rows:
            if row[nc_col - (0 if nc_col == 0 else 0)] and isinstance(row[nc_col - (0 if nc_col == 0 else 0)], str):
                campus = row[nc_col - (0 if nc_col == 0 else 0)]
                if 'Campus' in str(campus):
                    continue
                avg = row[c_col - (0 if c_col == 1 else 0)]
                # This logic is tricky -- the avg salary rows have Campus name in first col of each FY group
                pass

    # Handle average salaries separately -- they're in the last 3 rows
    avg_rows = [r for r in rows if r[0] and isinstance(r[0], str) and 'Ramirez' in str(r[0]) or (r[0] and isinstance(r[0], str) and 'Liberty' in str(r[0]))]
    for row in avg_rows:
        campus = str(row[0]).strip()
        for fy, avg_col in [('FY23', 1), ('FY24', 4), ('FY25', 7)]:
            avg_val = row[avg_col]
            if avg_val and isinstance(avg_val, (int, float)) and avg_val > 0:
                conn.execute("""
                    INSERT OR REPLACE INTO research_iltexas_campus_avg_salary
                    (source, campus_name, fiscal_year, avg_teacher_salary)
                    VALUES (?, ?, ?, ?)
                """, (SOURCE, campus, fy, float(avg_val)))
                avg_total += 1

    conn.commit()
    print(f"Loaded {salary_total} salary schedule rows, {avg_total} campus average salary rows")
    return salary_total + avg_total


def load_staffing(conn):
    """Load TA/para staffing from PIR 2.11.2026.xlsx S1 sheet."""
    print("\n=== Loading Staffing Data ===")

    wb = openpyxl.load_workbook(
        os.path.join(BASE_DIR, 'PIR 2.11.2026.xlsx'),
        data_only=True
    )

    ws = wb['S1. Campus - Level Staffing Det']
    rows = list(ws.iter_rows(values_only=True))

    total = 0
    school_year = None

    for row in rows:
        if row[0] and isinstance(row[0], str):
            # Check for year header
            if re.match(r'\d{4}-\d{4}', str(row[0]).strip()):
                school_year = str(row[0]).strip()
                continue

            campus = str(row[0]).strip()
            if campus in ('Campus', '') or school_year is None:
                continue

            ta_paras = int(row[1]) if row[1] and isinstance(row[1], (int, float)) else None
            ta_grade = int(row[2]) if row[2] and isinstance(row[2], (int, float)) else None
            cert_teachers = int(row[3]) if row[3] and isinstance(row[3], (int, float)) else None
            support_ftes = float(row[4]) if row[4] and isinstance(row[4], (int, float)) else None

            if ta_paras is not None:
                conn.execute("""
                    INSERT OR REPLACE INTO research_iltexas_staffing
                    (source, campus_name, school_year, total_ta_paras, ta_to_grade_level,
                     total_certified_teachers, total_support_staff_ftes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (SOURCE, campus, school_year, ta_paras, ta_grade,
                      cert_teachers, support_ftes))
                total += 1

    conn.commit()
    print(f"Loaded {total} staffing records")
    return total


def load_construction(conn):
    """Load Liberty County construction costs from CSV."""
    print("\n=== Loading Construction Costs ===")

    import csv
    csv_path = os.path.join(BASE_DIR, 'ILTexas Liberty County Construction Projects Summary 2.25.csv')

    total = 0
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            campus = row.get('Campus', '').strip()
            if not campus:
                continue

            conn.execute("""
                INSERT OR REPLACE INTO research_iltexas_construction
                (source, campus_name, construction_cost, general_contractor, architect,
                 sq_ft, design_enrollment, arch_engineering_cost, ffe_cost,
                 site_work_cost, inspectors_cost, land_acquisition_cost, capital_budget, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                SOURCE, campus,
                parse_currency(row.get('Construction Costs*', '')),
                row.get('General Contractor', '').strip(),
                row.get('Architect of Record', '').strip(),
                int(row.get('Sq Ft', '0').replace(',', '')) if row.get('Sq Ft', '').strip() else None,
                int(row.get('Enrollment', '0').replace(',', '')) if row.get('Enrollment', '').strip() else None,
                parse_currency(row.get('Architect/Engineering', '')),
                parse_currency(row.get('FFE', '')),
                parse_currency(row.get('Site Work (Clearing and Utilities) Outside GMP', '')),
                parse_currency(row.get('Inspectors', '')),
                parse_currency(row.get('Land Acquisition*', '')),
                row.get('Capital Budget', '').strip(),
                'All land was donated per ILTexas'
            ))
            total += 1

    conn.commit()
    print(f"Loaded {total} construction records")
    return total


def main():
    conn = get_db()
    create_tables(conn)

    results = {}
    results['teacher_counts'] = load_teacher_counts(conn)
    results['salary'] = load_salary_schedules(conn)
    results['staffing'] = load_staffing(conn)
    results['construction'] = load_construction(conn)
    results['peims'] = load_peims_campus_data(conn)

    conn.close()

    print("\n=== SUMMARY ===")
    for k, v in results.items():
        print(f"  {k}: {v} records")
    print(f"  TOTAL: {sum(results.values())} records")


if __name__ == '__main__':
    main()
